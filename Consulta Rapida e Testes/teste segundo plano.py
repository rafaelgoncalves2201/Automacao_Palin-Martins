from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import PatternFill
import openpyxl
import time
import datetime
from tkinter import *

servico = Service(ChromeDriverManager().install())
opcoes = webdriver.ChromeOptions()
opcoes.add_argument('--headless=new')
# Inicializar o WebDriver
driver = webdriver.Chrome(service=servico, options=opcoes)
driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')

# Carregar a planilha Excel
workbook = openpyxl.load_workbook('aiims.xlsx')  # Substitua 'aiims.xlsx' pelo nome do seu arquivo Excel
sheet = workbook.active

linha_planilha = 2  # Começar na segunda linha, supondo que a primeira linha seja cabeçalho

# Definir tempo de espera máximo
time.sleep(0.5)
wait = WebDriverWait(driver, 0.5)

# Define data de quando foi utilizado 
DATE = datetime.date.today().strftime("%d/%m/%Y")

cor_clickup = PatternFill(patternType='solid', fgColor='F0D402')
cor_outros = PatternFill(patternType='solid', fgColor='FF5B5B')
cor_naotem = PatternFill(patternType='solid', fgColor='55A3F9')

aiims_valido = {"Notificação do AIIM"}

outros = {
        "LITORAL", "OSASCO", 
        "CAPITAL I", "CAPITAL II", "CAPITAL III", 
        "GUARULHOS", 
        "DTE-II – FISCALIZAÇÃO ESPECIAL", "DTE-I – FISCALIZAÇÃO ESPECIAL",
        "Compliance MNM", "Compliance M&E"
        }


nomeColunas = ["N°", "DRT", "D.AIIM", 
            "CONTRIBUINTE", "ANDAMENTO DO AIIM",
            "CNPJ",
            "TELEFONE", "E-MAIL",
            "CNAE", "D.DIA", "SITUAÇÂO"
            ]

try:
    for col, nomeColunas in enumerate(nomeColunas, start=1):
        sheet.cell(row=1, column=col).value = nomeColunas
    # Loop pelas células com dados na planilha
    
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        aiim = row[0]
        if aiim is None or aiim == "":
            continue
        
        sheet.cell(row=linha_planilha, column=2).value = "DRT"

        # Espere até que o campo de AIIM seja clicável e insira o valor
        aiim_input = wait.until(EC.element_to_be_clickable((By.NAME, 'ctl00$ConteudoPagina$TxtNumAIIM')))
        aiim_input.clear()  # Limpa o campo antes de inserir o próximo AIIM
        aiim_input.send_keys(str(aiim))

        # Espere até que o botão de pesquisa seja clicável e clique nele
        pesquisar = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@title='Clique para pesquisar por numero do aiim (sem o digito verificador)']")))
        pesquisar.click()
        
        try:
            alert_wait = WebDriverWait(driver, timeout=0.5)
            alert = alert_wait.until(EC.alert_is_present())
            driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')
        
        except Exception as e:
            ERRO = "" # String Vazia
            DRT = ""
            
            try:
                # Espera até que o elemento com o ID 'ConteudoPagina_lblDRT' seja visível e obtenha o texto
                elemento_drt = wait.until(EC.visibility_of_element_located((By.ID, 'ConteudoPagina_lblDRT')))
                DRT = elemento_drt.text

                # Pesquisar nome
                elemento_nome = driver.find_element(By.ID, 'ConteudoPagina_lblNomeAutuado')
                NOME = elemento_nome.text

                # Pesquisar data
                elemento_data = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="dataEvento"]')))
                if elemento_data:
                    DATA = elemento_data[-1].text
                
                # Pesquisar descrição do andamento
                elemento_desc = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="descricaoEvento"]')))
                if elemento_desc:
                    DESC = elemento_desc[-1].text
                
            except Exception as e:
                # Se ocorrer uma exceção durante a busca de informações, registre o erro na planilha
                ERRO = str(e)
            
            finally:
                # Exibe a data do Dia
                sheet.cell(row=linha_planilha, column=10).value = DATE
                # Verificar se há erro e escrever na planilha
                
                if ERRO == "":
                    sheet.cell(row=linha_planilha, column=2).value = DRT
                    sheet.cell(row=linha_planilha, column=3).value = DATA
                    sheet.cell(row=linha_planilha, column=4).value = NOME
                    sheet.cell(row=linha_planilha, column=5).value = DESC
                    sheet.cell(row=linha_planilha, column=11).value = "Passar Click Up"
                    if DESC in aiims_valido:
                        sheet.cell(row=linha_planilha, column=11).value = "Passar ClickUp"
                        for col in range(1, 11):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_clickup
                    else:
                        sheet.cell(row=linha_planilha, column=11).value = "Suspenso"
                        for col in range(1, 11):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_outros
                
                else:
                    sheet.cell(row=linha_planilha, column=2).value = "Erro"
                    sheet.cell(row=linha_planilha, column=11).value = "Não tem ainda"
                    
                    for col in range(1, 11):
                        sheet.cell(row=linha_planilha, column=col).fill = cor_naotem
                    
                # Verificar se DRT está na lista outros
                if DRT in outros:
                    sheet.cell(row=linha_planilha, column=11).value = "Outros"
                    
                    for col in range(1, 11):
                        sheet.cell(row=linha_planilha, column=col).fill = cor_outros
                    

                # Incrementar o contador de linha em qualquer caso
                linha_planilha += 1

                # Volte para a página inicial após cada iteração
                driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')
            

finally:
    # Salvar o arquivo Excel
    workbook.save('Aiims Coletados.xlsx')  # Salvar com um novo nome para evitar a substituição do original

    # Fechar o navegador após o uso
    driver.quit()

    # Enviar mensagem de êxito
    print("Processo concluído com êxito!")
    
    
