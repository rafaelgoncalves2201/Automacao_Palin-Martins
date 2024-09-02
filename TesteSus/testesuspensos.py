import pyautogui
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl.styles import PatternFill
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import time
import openpyxl
import os
import glob
import shutil

# Configurações do WebDriver
servico = Service(ChromeDriverManager().install())

opcoes = webdriver.ChromeOptions()
opcoes.add_argument('--headless=new')
driver = webdriver.Chrome(service=servico, options=opcoes)

# Logar no ClickUp
driver.get("link")
print("Acessando o CRM para facilitar as buscas")

driver.find_element('xpath', '//*[@id="login-email-input"]').send_keys('exemplo')
driver.find_element('xpath', '//*[@id="login-password-input"]').send_keys('exemplo')
time.sleep(5)
# Sair di add
driver.find_element('xpath', '//*[@id="app-root"]/cu-login/div/div[2]/div[2]/div[1]/cu-login-form/div/form/button').click()
time.sleep(20)
# Clicar no filtro
driver.find_element('xpath', '//*[@id="app-root"]/cu-modal-keeper/cu-modal/div/div[2]/div[2]/div/cu-nux-modal/div/div[2]/button[1]').click()
time.sleep(5)
# Clicar para escolher filtro
driver.find_element('xpath', '//*[@id="app-root"]/cu-app-shell/cu-manager/div[1]/div/div/main/cu-dashboard/div/cu-views-bar-container/cu2-views-bar/div[1]/div[2]/div/cu-filter-block[1]/div/cu-filter-value-list-dropdown/div/div/div/span').click()
time.sleep(4)
# Clicar no Status
driver.find_element('xpath', '//*[@id="cdk-overlay-1"]/div/div/cu-search-list/cdk-virtual-scroll-viewport/div[1]/div/div[1]').click()
time.sleep(3)
# Clica na opção
driver.find_element('xpath', '//*[@id="cdk-overlay-0"]/div/div/cu-filter-value-list/div[1]/cu-filter-value-list/cu-filter-value/div/cu-status-filter/div/div/div[2]').click()
time.sleep(3)
# Escrever "Em andamento"
driver.find_element('xpath', '//*[@id="cdk-overlay-2"]/div/div[1]/div[1]/input').send_keys('Em Andamento')
time.sleep(4)
# Clicar na opção
driver.find_element('xpath', '//*[@id="cdk-overlay-2"]/div/div[2]/cu-status-filter-type-block/div[2]/div[1]/span[2]').click()
time.sleep(3)
# Clicar em confirmar
driver.find_element('xpath', '//*[@id="cdk-overlay-2"]/div/div[3]/div').click()
# Sair do filtro
pyautogui.press('esc')
time.sleep(5)
# Abrir a engrenagem
driver.find_element('xpath', '//*[@id="app-root"]/cu-app-shell/cu-manager/div[1]/div/div/main/cu-dashboard/div/cu-views-bar-container/cu2-views-bar/div[1]/div[1]/div/button[2]').click()
time.sleep(6)
# Clicar no Importar
driver.find_element('xpath', '//*[@id="app-root"]/cu-app-shell/cu-manager/div[1]/div/div/main/cu-dashboard/div/cu-views-bar-container/cu2-views-bar/div[3]/div/cu-dropdown-list-item[4]/button/div/div/cu-export-view/div/div/div').click()
time.sleep(6)
# Escolhe a opção de apenas nomes da tarefa
driver.find_element('xpath', '//*[@id="cdk-overlay-3"]/div/cu-dropdown-list-item[2]').click()
# Escolhe a opção excel
driver.find_element('xpath', '//*[@id="cdk-overlay-3"]/div/cu-dropdown-list-item[5]').click()
# Clica em baixar
driver.find_element('xpath', '//*[@id="cdk-overlay-3"]/div/button/div[1]').click()
time.sleep(30)
driver.quit()

# Gerenciar arquivos baixados
downloads = '/Users/Rafael/Downloads'
arquivo_xlsx = glob.glob(os.path.join(downloads, '*.xlsx'))
arquivo_recente = max(arquivo_xlsx, key=os.path.getctime)
diretorio = os.path.dirname(os.path.abspath(__file__))
destino_path = os.path.join(diretorio, "aiims desordenados.xlsx")
shutil.move(arquivo_recente, destino_path)

workbook = openpyxl.load_workbook('aiims desordenados.xlsx')
# Seleciona a planilha ativa
sheet = workbook.active
# Exclui as primeiras quatro linhas
sheet.delete_rows(1, 4)
# Insere uma nova linha no início da planilha
sheet.insert_rows(1)
# Salva as alterações no arquivo Excel
workbook.save('aiims susp.xlsx')


def formatar_numero(numero):
    numero = numero.replace('.', '').replace('-', '').replace(',', '')
    numero = numero[:7]
    return numero

servico = Service(ChromeDriverManager().install())

opcoes = webdriver.ChromeOptions()
opcoes.add_argument('--headless=new')


# Inicializar o WebDriver
driver = webdriver.Chrome(service=servico, options=opcoes)
driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')

print("Processo de consulta sendo realizada")

# Carregar a planilha Excel
workbook = openpyxl.load_workbook('aiims susp.xlsx')  # Substitua 'aiims.xlsx' pelo nome do seu arquivo Excel
sheet = workbook.active

linha_planilha = 2  # Começar na segunda linha, supondo que a primeira linha seja cabeçalho

# Definir tempo de espera máximo
time.sleep(0.5)
wait = WebDriverWait(driver, 0.5)

# Define data de quando foi utilizado
DATE = datetime.today()

cor_clickup = PatternFill(patternType='solid', fgColor='F0D402')
cor_outros = PatternFill(patternType='solid', fgColor='FF5B5B')
cor_naotem = PatternFill(patternType='solid', fgColor='55A3F9')

aiims_valido = {"Notificação do AIIM"
                
                }

aiims_verifi = {"Decurso de Prazo"
                }

aiims_invalido = {"Inscrição na Dívida Ativa/ AIIM inscrito em dívida ativa",
                "AIIM liquidado",
                "AIIM enviado para a Unidade Fiscal da Cobrança.",
                "Protocolo da Defesa",
                "Entrada do processo na Delegacia Tributária de Julgamento.",
                "Publicação no Diário Eletrônico",
                "Distribuição da Defesa para Julgamento",
                "Protocolo de Petição"
                }

nomeColunas = ["N°", "DRT", "D.AIIM",
               "CONTRIBUINTE", "CNPJ",
               "TELEFONE", "E-MAIL",
               "CNAE", "D.DIA", "SITUAÇÂO"
               ]

try:
    for col, nomeColunas in enumerate(nomeColunas, start=1):
        sheet.cell(row=1, column=col).value = nomeColunas

    # Loop pelas células com dados na planilha
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        aiim = row[0]
        if aiim is None or aiim == "":
            continue

        aiim_formatado = formatar_numero(aiim)

        sheet.cell(row=linha_planilha, column=2).value = "DRT"

        # Espere até que o campo de AIIM seja clicável e insira o valor
        aiim_input = wait.until(EC.element_to_be_clickable((By.NAME, 'ctl00$ConteudoPagina$TxtNumAIIM')))
        aiim_input.clear()  # Limpa o campo antes de inserir o próximo AIIM
        aiim_input.send_keys(aiim_formatado)

        # Espere até que o botão de pesquisa seja clicável e clique nele
        pesquisar = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@title='Clique para pesquisar por numero do aiim (sem o digito verificador)']")))
        pesquisar.click()

        try:
            alert_wait = WebDriverWait(driver, timeout=0.5)
            alert = alert_wait.until(EC.alert_is_present())
            driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')

        except Exception as e:
            ERRO = "" # String Vazia
            DRT = ""

            try:
                # Espera até que o elemento com o ID 'ConteudoPagina_lblDRT' seja visível e obtenha o texto
                elemento_drt = wait.until(EC.visibility_of_element_located((By.ID, 'ConteudoPagina_lblDRT')))
                DRT = elemento_drt.text

                # Pesquisar nome
                elemento_nome = driver.find_element(By.ID, 'ConteudoPagina_lblNomeAutuado')
                NOME = elemento_nome.text

                # Pesquisar data
                elemento_data = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="dataEvento"]')))
                if elemento_data:
                    DATA = datetime.strptime(elemento_data[-1].text, "%d/%m/%Y")

                # Pesquisar descrição do andamento
                elemento_desc = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="descricaoEvento"]')))
                if elemento_desc:
                    DESC = elemento_desc[-1].text

            except Exception as e:
                # Se ocorrer uma exceção durante a busca de informações, registre o erro na planilha
                ERRO = str(e)

            finally:
                # Exibe a data do Dia
                sheet.cell(row=linha_planilha, column=9).value = DATE
                # Verificar se há erro e escrever na planilha

                if ERRO == "":
                    sheet.cell(row=linha_planilha, column=10).value = DATE.strftime("%d/%m/%Y")
                    
                # Verificar se há erro e escrever na planilha
                if ERRO == "":
                    sheet.cell(row=linha_planilha, column=2).value = DRT
                    sheet.cell(row=linha_planilha, column=3).value = DATA.strftime("%d/%m/%Y")
                    sheet.cell(row=linha_planilha, column=4).value = NOME
                    sheet.cell(row=linha_planilha, column=5).value = DESC
                    
                    def diferenca_boa(DATE, DATA):
                    
                        d1 = DATE
                        d2 = DATA
                        diferenca_a = abs(( d2 - d1 ).days)
                        
                        return diferenca_a < 90
                    
                    # Comparação para tomada de decisão
                    if DESC in aiims_valido:
                        sheet.cell(row=linha_planilha, column=11).value = "Passar ClickUp"
                        for col in range(1, 12):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_clickup
                    elif DESC in aiims_invalido:
                        sheet.cell(row=linha_planilha, column=11).value = "Suspenso"
                        for col in range(1, 12):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_outros
                    elif DESC in aiims_verifi:
                        sheet.cell(row=linha_planilha, column=11).value = "Passar ClickUp"
                        for col in range(1, 12):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_clickup
                    elif DATA and diferenca_boa(DATE, DATA):
                        sheet.cell(row=linha_planilha, column=11).value = "Passar ClickUp"
                        for col in range(1, 12):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_clickup
                    else:
                        sheet.cell(row=linha_planilha, column=11).value = "Suspenso"
                        for col in range(1, 12):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_outros
                            
                        
                # Incrementar o contador de linha em qualquer caso
                linha_planilha += 1

                # Volte para a página inicial após cada iteração
                driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')


finally:
    # Salvar o arquivo Excel
    workbook.save('Verificação de Aiims Suspenso.xlsx')  # Salvar com um novo nome para evitar a substituição do original

    # Fechar o navegador após o uso
    driver.quit()

    # Enviar mensagem de êxito
    print("Processo concluído com êxito!")
