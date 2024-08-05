import os
import sys
import threading
import re
from datetime import datetime
from tkinter import ttk, filedialog, Tk, Label, Frame, Button
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# Carregar ícone
def carregar_icone():
    base_path = getattr(sys, 'frozen', False) and sys._MEIPASS or os.path.dirname(__file__)
    return os.path.join(base_path, "imgs", "palinico.ico")

def normalizar_descricao(desc):
    padroes = {
        r'\bDecurso de Prazo\b.*': 'Decurso de Prazo',
        r'\bDistribuição da Defesa para Julgamento\b.*': 'Distribuição da Defesa para Julgamento',
        r'\bProtocolo de Petição\b.*': 'Protocolo de Petição'
    }
    for padrao, substituicao in padroes.items():
        if re.search(padrao, desc):
            return re.sub(padrao, substituicao, desc).strip()
    return desc.strip()

def formatar_aiim(aiim):
    apenas_numeros = re.sub(r'\D', '', aiim)
    return apenas_numeros[:7] + apenas_numeros[8:] if len(apenas_numeros) >= 8 else apenas_numeros

def exibir(tipo, texto):
    mensagem_label.config(text=texto, fg="black")

def escolher_planilha():
    global file_path
    file_path = filedialog.askopenfilename(title="Selecione a planilha Excel", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        exibir("Sucesso", "Planilha selecionada com sucesso.")

def mostrar_progress_bar():
    progress_bar.pack(pady=20)
    janela.update_idletasks()

def esconder_progress_bar():
    progress_bar.pack_forget()

def atualizar_progress_bar(valor):
    progress_bar['value'] = valor
    janela.update_idletasks()
    
def cancelar_processo_funcao():
    global cancelar_processo
    cancelar_processo = True
    exibir("Cancelar", "Processo cancelado.")
    esconder_progress_bar()
    if driver:
        driver.quit()

def iniciar_processo():
    if not file_path:
        exibir("Erro", "Por favor, selecione uma planilha primeiro.")
        return

    exibir("Processo", "Processo iniciado, por favor aguarde...")
    mostrar_progress_bar()

    threading.Thread(target=processo_automacao).start()
    
def navegador():
    global driver
    if driver is None:
        servico = Service(ChromeDriverManager().install())
        opcoes = webdriver.ChromeOptions()
        opcoes.add_argument('--headless')
        opcoes.add_argument('--disable-gpu')
        opcoes.add_argument('--no-sandbox')
        opcoes.add_argument('--disable-dev-shm-usage')
        driver = webdriver.Chrome(service=servico, options=opcoes)
    return driver

def salvar_planilha(workbook):
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Salvar planilha como")
    if save_path:
        workbook.save(save_path)
        exibir("Pronto", "Planilha salva com sucesso.")
        janela.event_generate("<<ProcessoSalvo>>")

def processo_automacao():
    global driver, file_path, cancelar_processo

    if not file_path:
        return
    
    driver = navegador()

    try:
        driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')

        workbook = load_workbook(file_path)
        sheet = workbook.active
        linha_planilha = 2
        wait = WebDriverWait(driver, 0.5)

        DATE = datetime.today()
        cor_clickup = PatternFill(patternType='solid', fgColor='F0D402')
        cor_outros = PatternFill(patternType='solid', fgColor='FF5B5B')
        cor_naotem = PatternFill(patternType='solid', fgColor='55A3F9')

        aiims_valido = {"Notificação do AIIM",
                        "Inscrição na Dívida Ativa/ AIIM inscrito em dívida ativa",
                        "AIIM enviado para a Unidade Fiscal da Cobrança.", "Decurso de Prazo",
                        "Ratificação do AIIM"}
        
        aiims_invalido = {"AIIM liquidado", "Protocolo da Defesa", "Protocolo de Petição",
                          "Entrada do processo na Delegacia Tributária de Julgamento.",
                          "Publicação no Diário Eletrônico",
                          "Distribuição da Defesa para Julgamento", "Protocolo de Petição"}
        
        outros = {"LITORAL", "OSASCO", "CAPITAL I", "CAPITAL II", "CAPITAL III", "GUARULHOS",
                  "DTE-II – FISCALIZAÇÃO ESPECIAL", "DTE-I – FISCALIZAÇÃO ESPECIAL",
                  "Compliance MNM", "Compliance M&E"}

        nomeColunas = ["N°", "DRT", "D.AIIM", "CONTRIBUINTE", "ANDAMENTO DO AIIM",
                       "CNPJ", "TELEFONE", "E-MAIL", "CNAE", "D.DIA", "SITUAÇÂO"]

        for col, nomeColuna in enumerate(nomeColunas, start=1):
            sheet.cell(row=1, column=col).value = nomeColuna

        total_rows = len(list(sheet.iter_rows(min_row=2, max_col=1, values_only=True)))
        atualizar_progress_bar(0)  # Inicializa a barra de progresso com 0%

        for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=1):
            if cancelar_processo:
                exibir("Cancelar", "Processo cancelado.")
                break
            
            aiim = row[0]
            if aiim is None or aiim == "":
                continue

            sheet.cell(row=linha_planilha, column=2).value = "DRT"
            aiim_format = formatar_aiim(str(aiim))
            aiim_input = wait.until(EC.element_to_be_clickable((By.NAME, 'ctl00$ConteudoPagina$TxtNumAIIM')))
            aiim_input.clear()
            aiim_input.send_keys(aiim_format)

            pesquisar = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@title='Clique para pesquisar por numero do aiim (sem o digito verificador)']")))
            pesquisar.click()

            try:
                alert_wait = WebDriverWait(driver, timeout=0.5)
                alert = alert_wait.until(EC.alert_is_present())
                driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')
                continue

            except Exception:
                ERRO = ""
                DRT = ""
                NOME = ""
                DATA = ""
                DESC = ""
                try:
                    elemento_drt = wait.until(EC.visibility_of_element_located((By.ID, 'ConteudoPagina_lblDRT')))
                    DRT = elemento_drt.text

                    elemento_nome = driver.find_element(By.ID, 'ConteudoPagina_lblNomeAutuado')
                    NOME = elemento_nome.text

                    elemento_data = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="dataEvento"]')))
                    if elemento_data:
                        DATA = datetime.strptime(elemento_data[-1].text, "%d/%m/%Y")

                    elemento_desc = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="descricaoEvento"]')))
                    
                    ultima_situacao = None
                    for elemento in elemento_desc:
                        descricao = elemento.text
                        ultima_situacao = descricao
                        if any(invalido in descricao for invalido in aiims_invalido):
                            break
                        
                    DESC = ultima_situacao
                    
                except Exception as e:
                    ERRO = str(e)

                finally:
                    sheet.cell(row=linha_planilha, column=10).value = DATE.strftime("%d/%m/%Y")

                    desc = normalizar_descricao(DESC)
                    if ERRO == "":
                        sheet.cell(row=linha_planilha, column=2).value = DRT
                        sheet.cell(row=linha_planilha, column=3).value = DATA.strftime("%d/%m/%Y")
                        sheet.cell(row=linha_planilha, column=4).value = NOME
                        sheet.cell(row=linha_planilha, column=5).value = DESC
                        
                        if desc in aiims_valido:
                            sheet.cell(row=linha_planilha, column=11).value = "Passar ClickUp"
                            for col in range(1, 12):
                                sheet.cell(row=linha_planilha, column=col).fill = cor_clickup
                        else:
                            sheet.cell(row=linha_planilha, column=11).value = "Suspenso"
                            for col in range(1, 12):
                                sheet.cell(row=linha_planilha, column=col).fill = cor_outros
                                
                        if desc in aiims_invalido:
                            sheet.cell(row=linha_planilha, column=11).value = "Suspenso"
                            for col in range(1, 12):
                                sheet.cell(row=linha_planilha, column=col).fill = cor_outros
                                
                        if DRT in outros:
                            sheet.cell(row=linha_planilha, column=11).value = "Outros"
                            for col in range(1, 12):
                                sheet.cell(row=linha_planilha, column=col).fill = cor_outros
                                
                    else:
                        sheet.cell(row=linha_planilha, column=2).value = "Erro"
                        sheet.cell(row=linha_planilha, column=11).value = "Não tem ainda"
                        for col in range(1, 12):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_naotem

                    driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')

                    atualizar_progress_bar((index / total_rows) * 100)
                    linha_planilha += 1
                    
    except Exception as e:
        janela.event_generate("<<ProcessoErro>>", data=str(e))
              
    finally:
        salvar_planilha(workbook)    
        esconder_progress_bar()
        if driver:
            driver.quit()


def atualizar_status(event):
    exibir("Processo finalizado ", "Salve a planilha")
    
def atualizar_salvamento(event):
    exibir("Pronto", "Tudo Salvo")

def mostrar_erro(event):
    exibir("Erro", f"Erro inesperado: {event.data}")
    
def nova_consulta():
    global file_path, cancelar_processo, driver
    
    file_path = None
    cancelar_processo = True
    
    if driver:
        driver.quit()
        driver = None
    
    exibir("Calma", "Nova consulta iniciada. Adicione uma nova planilha.")
    mensagem_label.config(text="", fg="black")
    atualizar_progress_bar(0)
    
    progress_bar['value'] = 0
    janela.update_idletasks()
    
# Criação da janela principal
janela = Tk()
janela.title("Automação para Consulta de Aiims")
janela.geometry("800x600")
janela.configure(bg="#f8f9fa")

# Carregar ícone
def carregar_icone():
    if getattr(sys, 'frozen', False):
        icone_path = os.path.join(sys._MEIPASS, "imgs", "palinico.ico")
    else:
        icone_path = os.path.join("imgs", "palinico.ico")
    return icone_path

icone_path = carregar_icone()
janela.iconbitmap(icone_path)

header_frame = Frame(janela, bg="#007bff", padx=20, pady=10)
header_frame.pack(fill="x")

header_label = Label(header_frame, text="Automação para Consulta de Aiims", bg="#007bff", fg="white", font=("Helvetica", 16, "bold"))
header_label.pack()

body_frame = Frame(janela, bg="#f8f9fa")
body_frame.pack(padx=20, pady=20, fill="both", expand=True)

instruction_label = Label(body_frame, text="Por favor, adicione a planilha Excel contendo os Aiims:", bg="#f8f9fa", fg="#495057", font=("Helvetica", 12))
instruction_label.pack(pady=10)

button_frame = Frame(body_frame, bg="#f8f9fa")
button_frame.pack(pady=20)

planilha_btn = Button(button_frame, text="Adicionar Planilha", command=escolher_planilha, bg="#007bff", fg="white", font=("Helvetica", 12), relief="flat", padx=10, pady=5)
planilha_btn.pack(pady=10)

iniciar_btn = Button(button_frame, text="Iniciar", command=iniciar_processo, bg="#007bff", fg="white", font=("Helvetica", 12), relief="flat", padx=10, pady=5)
iniciar_btn.pack(pady=10)

nova_consulta_btn = Button(janela, text="Iniciar nova consulta", bg="#007bff", fg="white", font=("Helvetica", 12), relief="flat", padx=10, pady=5)
nova_consulta_btn.pack(pady=10)

cancelar_btn = Button(janela, text="Cancelar", command=cancelar_processo_funcao, bg="#dc3545", fg="white", font=("Helvetica", 12), relief="flat", padx=10, pady=5)
cancelar_btn.pack(pady=10)

mensagem_label = Label(body_frame, text="", bg="#f8f9fa", font=("Helvetica", 12))
mensagem_label.pack(pady=10)

# Barra de Progresso
progress_bar = ttk.Progressbar(body_frame, orient="horizontal", length=400, mode="determinate")

file_path = None
cancelar_processo = False
driver = None

def atualizar_nova_consulta(event):
    exibir("Nova consulta", "Adicione uma nova planilha.")
    atualizar_progress_bar(0)
    
janela.bind("<<NovaConsultaIniciada>>", atualizar_nova_consulta)
janela.bind("<<ProcessoConcluido>>", atualizar_status)
janela.bind("<<ProcessoSalvo>>", atualizar_salvamento)
janela.bind("<<ProcessoErro>>", mostrar_erro)

janela.mainloop()
# pyinstaller --onefile --windowed --add-data "imgs/palinico.ico;imgs" --icon=imgs/palinico.ico Consulta.py
