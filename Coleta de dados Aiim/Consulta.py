from tkinter import ttk
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time

def exibir(tipo, texto):
    cor = "white" if tipo == "Sucesso" else "white"
    mensagem_label.config(text=texto, fg=cor)

def escolher_planilha():
    global file_path
    file_path = filedialog.askopenfilename(title="Selecione a planilha Excel", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        exibir("Sucesso", "Planilha selecionada com sucesso.")

def executar_processo():
    global file_path
    if not file_path:
        exibir("Erro", "Por favor, selecione uma planilha primeiro.")
        progresso.stop()
        return

    progresso.start()

    try:
        exibir("Processo", "processo iniciado, por favor aguarde...")
        servico = Service(ChromeDriverManager().install())
        opcoes = webdriver.ChromeOptions()
        opcoes.add_argument('--headless=new')
        driver = webdriver.Chrome(service=servico, options=opcoes)
        driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')

        workbook = load_workbook(file_path)
        sheet = workbook.active

        linha_planilha = 2
        time.sleep(0.5)
        wait = WebDriverWait(driver, 0.5)

        DATE = datetime.today()

        cor_clickup = PatternFill(patternType='solid', fgColor='F0D402')
        cor_outros = PatternFill(patternType='solid', fgColor='FF5B5B')
        cor_naotem = PatternFill(patternType='solid', fgColor='55A3F9')

        aiims_valido = {"Notificação do AIIM"}

        aiims_verifi = {"Decurso de Prazo"
                        }

        aiims_invalido = {"Inscrição na Dívida Ativa/ AIIM inscrito em dívida ativa",
                        "AIIM liquidado",
                        "AIIM enviado para a Unidade Fiscal da Cobrança.",
                        "Protocolo da Defesa",
                        "Entrada do processo na Delegacia Tributária de Julgamento.",
                        "Publicação no Diário Eletrônico",
                        "Distribuição da Defesa para Julgamento",
                        "Protocolo de Petição"
                        }

        outros = {
            "LITORAL", "OSASCO", 
            "CAPITAL I", "CAPITAL II", "CAPITAL III", 
            "GUARULHOS", 
            "DTE-II – FISCALIZAÇÃO ESPECIAL", "DTE-I – FISCALIZAÇÃO ESPECIAL",
            "Compliance MNM", "Compliance M&E"
        }

        nomeColunas = ["N°", "DRT", "D.AIIM", "CONTRIBUINTE", "ANDAMENTO DO AIIM", "CNPJ", "TELEFONE", "E-MAIL", "CNAE", "D.DIA", "SITUAÇÂO"]

        for col, nomeColuna in enumerate(nomeColunas, start=1):
            sheet.cell(row=1, column=col).value = nomeColuna

        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            aiim = row[0]
            if aiim is None or aiim == "":
                continue

            sheet.cell(row=linha_planilha, column=2).value = "DRT"

            aiim_input = wait.until(EC.element_to_be_clickable((By.NAME, 'ctl00$ConteudoPagina$TxtNumAIIM')))
            aiim_input.clear()
            aiim_input.send_keys(str(aiim))

            pesquisar = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@title='Clique para pesquisar por numero do aiim (sem o digito verificador)']")))
            pesquisar.click()

            try:
                alert_wait = WebDriverWait(driver, timeout=0.5)
                alerat = alert_wait.until(EC.alert_is_present())
                driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')

            except Exception as e:
                ERRO = ""
                DRT = ""

                try:
                    elemento_drt = wait.until(EC.visibility_of_element_located((By.ID, 'ConteudoPagina_lblDRT')))
                    DRT = elemento_drt.text

                    elemento_nome = driver.find_element(By.ID, 'ConteudoPagina_lblNomeAutuado')
                    NOME = elemento_nome.text

                    elemento_data = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="dataEvento"]')))
                    if elemento_data:
                        DATA = datetime.strptime(elemento_data[-1].text, "%d/%m/%Y")

                    elemento_desc = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="descricaoEvento"]')))
                    if elemento_desc:
                        DESC = elemento_desc[-1].text

                except Exception as e:
                    ERRO = str(e)
                # Verificar se há erro e escrever na planilha

                finally:
                    sheet.cell(row=linha_planilha, column=10).value = DATE.strftime("%d/%m/%Y")
                    
                # Verificar se há erro e escrever na planilha
                if ERRO == "":
                    sheet.cell(row=linha_planilha, column=2).value = DRT
                    sheet.cell(row=linha_planilha, column=3).value = DATA.strftime("%d/%m/%Y")
                    sheet.cell(row=linha_planilha, column=4).value = NOME
                    sheet.cell(row=linha_planilha, column=5).value = DESC
                    
                    def diferenca_boa(DATE, DATA):
                    
                        d1 = DATE
                        d2 = DATA
                        diferenca_a = abs(( d2 - d1 ).days)
                        
                        return diferenca_a < 90
                    
                    # Comparação para tomada de decisão
                    if DESC in aiims_valido:
                        sheet.cell(row=linha_planilha, column=11).value = "Passar ClickUp"
                        for col in range(1, 12):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_clickup
                        if DESC in aiims_invalido:
                            sheet.cell(row=linha_planilha, column=11).value = "Suspenso"
                            for col in range(1, 12):
                                sheet.cell(row=linha_planilha, column=col).fill = cor_outros
                        elif DESC in aiims_verifi:
                            sheet.cell(row=linha_planilha, column=11).value = "Passar ClickUp"
                            for col in range(1, 12):
                                sheet.cell(row=linha_planilha, column=col).fill = cor_clickup
                        elif DATA and diferenca_boa(DATE, DATA):
                            sheet.cell(row=linha_planilha, column=11).value = "Passar ClickUp"
                            for col in range(1, 12):
                                sheet.cell(row=linha_planilha, column=col).fill = cor_clickup
                        else:
                            sheet.cell(row=linha_planilha, column=11).value = "Suspenso"
                            for col in range(1, 12):
                                sheet.cell(row=linha_planilha, column=col).fill = cor_outros

                    else:
                        sheet.cell(row=linha_planilha, column=2).value = "Erro"
                        sheet.cell(row=linha_planilha, column=11).value = "Não tem ainda"

                        for col in range(1, 12):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_naotem

                    if DRT in outros:
                        sheet.cell(row=linha_planilha, column=11).value = "Outros"

                        for col in range(1, 12):
                            sheet.cell(row=linha_planilha, column=col).fill = cor_outros

                    linha_planilha += 1
                    driver.get('https://www.fazenda.sp.gov.br/epat/extratoprocesso/PesquisarExtrato.aspx')
        exibir("Pronto", "Agora salve a planilha ja consultada!")
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Salvar planilha como")
        if save_path:
            workbook.save(save_path)
            exibir("Sucesso", "Processo concluído com êxito!")
        progresso.stop()
        driver.quit()
    except Exception as e:
        exibir("Erro", str(e))
        progresso.stop()

janela = tk.Tk()
janela.title("Automação para consulta de Aiims")
janela.geometry("720x640")
janela.configure(background="#1a1a1a")

label = tk.Label(janela, text="Por favor, adicione a planilha Excel contendo os Aiims:", bg="#1a1a1a", fg="white", font=("Helvetica", 12))
label.pack(pady=20)

frame = tk.Frame(janela, bg="#1a1a1a")
frame.pack(pady=20)

planilha = tk.Button(frame, text="Adicionar Planilha", command=escolher_planilha, bg="#e3c13e", fg="#1a1a1a", font=("Helvetica", 12))
planilha.pack(pady=10)

iniciar = tk.Button(frame, text="Iniciar", command=executar_processo, bg="#e3c13e", fg="#1a1a1a", font=("Helvetica", 12))
iniciar.pack(pady=10)

mensagem_label = tk.Label(janela, text="", bg="#1a1a1a", font=("Helvetica", 12))
mensagem_label.pack(pady=10)

progresso = ttk.Progressbar(janela, orient="horizontal", length=300, mode="determinate")
progresso.pack(pady=20)

file_path = None

janela.mainloop()

# Buildar pyinstaller --onefile -w Consulta.py